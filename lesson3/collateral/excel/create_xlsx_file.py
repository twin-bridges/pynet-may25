"""
This is a more complex example where I read in a data structure from YAML
and then create an Excel spreadsheet from this data.
"""
import yaml
from rich import print
import openpyxl


def read_yaml_file(filename):
    with open(filename) as f:
        return yaml.safe_load(f)


if __name__ == "__main__":

    # Retrieve source YAML data
    yaml_data = read_yaml_file("netmiko.yml")
    print(yaml_data)

    new_wb = openpyxl.Workbook()
    sheet = new_wb.active
    sheet.title = "Net Devices"

    # Convert 'yaml_data' to a list of dictionaries
    devices = []
    for device_name, device_dict in yaml_data.items():
        device_dict["device_name"] = device_name
        devices.append(device_dict)

    header_row = [
        "device_name",
        "device_type",
        "host",
        "username",
        "password",
        "port",
        "global_delay_factor",
    ]

    # Create the header row, start at 1 (as cells are 1-indexed)
    for column, header_name in enumerate(header_row, start=1):
        # Access 'cell' using row and column numbers
        sheet.cell(row=1, column=column).value = header_name

    # Add the device data, start at row=2 as we already have the header row.
    for row, device_dict in enumerate(devices, start=2):
        for column, field_name in enumerate(header_row, start=1):
            # Some devices don't have all of the fields, set missing fields to null-string
            cell_value = device_dict.get(field_name, "")
            sheet.cell(row=row, column=column).value = cell_value

    new_wb.save("my_file.xlsx")

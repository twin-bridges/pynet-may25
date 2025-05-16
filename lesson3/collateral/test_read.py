import openpyxl

# Load the workbook
# wb = openpyxl.load_workbook('my_file.xlsx')
wb = openpyxl.load_workbook('net_devices_updated.xlsx')

# Get the active sheet
sheet = wb.active
for row in sheet.iter_rows(values_only=True):
    print(row)

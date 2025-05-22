"""
Using Python and the openpyxl library, open the following spreadsheet:

net_devices.xlsx

You should save this file to your lab server and access it using openpyxl.

Using openpyxl and the net_devices.xlsx spreadsheet change the name of "srx2" to "srx2_new"
(srx2 is in the "A8" cell).

Add a new network device named "srx1" to the spreadsheet. All of the other fields besides
the name and host fields should be set the same as "srx2". Note, you can add a row to the
spreedsheet by doing the following:

# new_netdevice is a list with all of the fields needed for the new row in the spreadsheet.
sheet.append(new_netdevice)

Save this updated spreadsheet to a new filename. You can close the workbook by calling
"workbook.close()" (after the save() operation).

As a verification step, read in the new spreadsheet you just saved (using .load_workbook())
and then verify the new name for "srx2" and also verify new row you added.

"""
import openpyxl

# Load the workbook
wb = openpyxl.load_workbook("net_devices.xlsx")

# Get the active sheet
sheet = wb.active

# Change the name for SRX2
srx_name = sheet["A8"]
srx_name.value = "srx2_new"

# Insert a completely new network device
new_netdevice = (
    "srx1",
    "juniper_junos",
    "srx1.invalid.com",
    "admin",
    "juniper123",
    "",
    "",
)
sheet.append(new_netdevice)

new_filename = "net_devices_text.xlsx"
wb.save(new_filename)
wb.close()

# Verification:
print("\n* Verification of new spreadsheet *\n")
wb = openpyxl.load_workbook(new_filename)
srx_name = sheet["A8"]
print(f"srx2 new name: {srx_name.value}")

print("\nVerify new 'srx1' row: ")
last_row = sheet["11"]
for field in last_row:
    if not field.value:
        # So empty fields/null-strings show up
        print("empty")
        continue
    print(field.value)
print()

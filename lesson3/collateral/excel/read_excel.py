from rich import print
import openpyxl

# import ipdb; ipdb.set_trace()
import pdbr

# Load the workbook
workbook = openpyxl.load_workbook("net_devices.xlsx")

# Get the active sheet
sheet = workbook.active

pdbr.set_trace()

# Access a cell directly
print(sheet["A1"])

# Cell value
print(sheet["A1"].value)

# Access the "A" column (returns a tuple)
print(sheet["A"])

# We could loop over this column:
for cell in sheet["A"]:
    print(cell.value)

# Access the 1st row (returns a tuple)
print(sheet["1"])

# We could loop over this row:
for cell in sheet["1"]:
    print(cell.value)

# Access multiple columns
print(sheet["D:G"])

# Access multiple rows
print(sheet["4:6"])

# Access multiple rows and columns (across first, then down)
print(sheet["A1:G2"])
print(sheet["A1:B3"])

# Can directly iterate over the rows.
for row in sheet.iter_rows():
    print(row)
    break

# Can directly iterate over the columns
for column in sheet.iter_cols():
    print(column)
    break

# Accessing a cell by number
cell = sheet.cell(row=1, column=1)
print(cell.value)

new_filename = "net_devices_updated.xlsx"
workbook.save(new_filename)
